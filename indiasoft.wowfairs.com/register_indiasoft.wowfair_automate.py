from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import time
#from selenium.webdriver.support.select import Select 
# import openpyxl module 
import openpyxl 
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook("C:\\Users\BP\\Desktop\\indiasoft.wowfairs.com\\new_foreign_delegates-2.xlsx") 

# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

# Cell object is created by using  
# sheet object's cell() method. 

driver_chrome = webdriver.Chrome("C:\chromedriver_win32\chromedriver.exe")
driver_chrome.get("https://indiasoft.wowfairs.com/")
# it will maximize the window size, bydefault the window size is not maximize
driver_chrome.maximize_window()


for i in range(1,19):
    
        alpha = sheet_obj.cell(row = i, column = 1)
        a = alpha.value
        print(a)
        b = sheet_obj.cell(row = i, column = 2)
        b = b. value
        c = sheet_obj.cell(row = i, column = 3)
        c = c.value
        d = sheet_obj.cell(row = i, column = 4)
        d = d.value
       
        
        def click():
            element.click()
            time.sleep(2) 
            
        #search for register butoon    
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/form/div[6]/button[1]")
        click()   # click on register button
        
        # full name
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[1]/input")
        element.send_keys(a)
        
        
        # email address
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[2]/input")
        element.send_keys(b)
        
        # password
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[3]/input")
        element.send_keys(c)
        
        # confirm password
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[4]/input")
        element.send_keys(d)
        
               
        # terms and condition check box
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[9]/label/input")
        click()
        
        #click on register
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/button")
        element.click()
        time.sleep(5)
        
        driver_chrome.get("https://indiasoft.wowfairs.com/")
        
        
       
        
        
        
        
        
        
        
        
        
        
        
