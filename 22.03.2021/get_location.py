from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import time
import requests
import json

#from selenium.webdriver.support.select import Select 
# import openpyxl module 
import openpyxl 
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook("D:\\address.xlsx") 

# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

# Cell object is created by using  
# sheet object's cell() method. 

for i in range(2,4):
        driver_chrome = webdriver.Chrome("C:\chromedriver_win32\chromedriver.exe")
        latitude = sheet_obj.cell(row = i, column = 1)
        a = latitude.value
        print(a)
        longitude = sheet_obj.cell(row = i, column = 2)
        b = longitude. value
        print(b)
        driver_chrome.get("https://maps.googleapis.com/maps/api/geocode/json?address={},{}&key=AIzaSyD4TRlWnZ6zqz4LsxNhPsZ87mZItOZ6D7w".format(a,b))
        time.sleep(3)
        driver_chrome.maximize_window()
        time.sleep(3)
        driver_chrome.close()
        time.sleep(3)
        
        





import requests
import json
import re


r = requests.get("https://maps.googleapis.com/maps/api/geocode/json?address=43.61288299935499,%20-107.51342285996284&key=AIzaSyD4TRlWnZ6zqz4LsxNhPsZ87mZItOZ6D7w")


#package_name = package_json[0]['name']
package_json = r.json()

# dumps will make json into string and  indentation as it shows on api
# Use the separators parameter to change the default separator:
    
package_str = json.dumps(package_json, indent = 2, separators=(". ", " = "))
# Use the sort_keys parameter to specify if the result should be sorted or not:
#package_str = json.dumps(package_json, indent=2, sort_keys=True)

#x = re.search("^long_name.*political$", package_str)

print(package_str)

