from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import time
#from selenium.webdriver.support.select import Select
import openpyxl 

wb_obj = openpyxl.load_workbook("C:\\Users\\user\\Desktop\\wow\\chat.wowfairs\\exhibitors_list.xlsx")
sheet_obj = wb_obj.active

user_name = "gaurav.raheja@outlook.com"
password = "asdf@1234"

driver_chrome = webdriver.Chrome("C:\chromedriver_win32\chromedriver.exe")
driver_chrome.get("https://chat.wowfairs.com/login")
driver_chrome.maximize_window()
# it will maximize the window size, bydefault the window size is not maximize

element = driver_chrome.find_element_by_xpath("/html/body/div/main/div/div/div/div/div[2]/form/div[1]/div/input")
element.send_keys(user_name)
# it will search for the element id inputPassword
element = driver_chrome.find_element_by_xpath("/html/body/div/main/div/div/div/div/div[2]/form/div[2]/div/input")
element.send_keys(password) # it will pass the string value assigned to password variable
time.sleep(2)
    
#click on remember me check box
element = driver_chrome.find_element_by_xpath("/html/body/div/main/div/div/div/div/div[2]/form/div[3]/div/label")
element.click()
# it will press the enter after values have been entered
element = driver_chrome.find_element_by_xpath("/html/body/div/main/div/div/div/div/div[2]/form/div[4]/div/button") 
element.click()
time.sleep(4)

# click on exhibitor management
element = driver_chrome.find_element_by_xpath("/html/body/div[4]/section/aside/nav/ul/li[2]/a")
element.click()
time.sleep(2)
# click on add new
element = driver_chrome.find_element_by_xpath("/html/body/div[4]/section/aside/nav/ul/li[2]/ul/li[2]/a")
element.click()
time.sleep(2)

for i in range(1,130):
    
        alpha = sheet_obj.cell(row = i, column = 1)
        a = alpha.value
        print(a)
        b = sheet_obj.cell(row = i, column = 2)
        b = b. value
        c = sheet_obj.cell(row = i, column = 3)
        c = c.value
        d = sheet_obj.cell(row = i, column = 4)
        d = d.value
        e = sheet_obj.cell(row = i, column = 5)
        e = e.value
  

        #name
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[1]/div/input")
        element.send_keys(a) 

        #email
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[2]/div/input")
        element.send_keys(b) # it will pass the string value assigned 

        #password
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[3]/div/input")
        element.send_keys(c) # it will pass the string value assigned 

        #confirm password
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[4]/div/input")
        element.send_keys(d) # it will pass the string value assigned 

        # department
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[6]/div/span/span[1]/span/span[1]")
        element.click() # it will pass the string value assigned 
        time.sleep(1)
        
        
        element = driver_chrome.find_element_by_xpath("/html/body/span/span/span[2]/ul/li[3]")
        element.click()


        #click on save
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/section/div[2]/div/div/div/div[2]/form/div[1]/div[7]/div/button[2]")
        element.click()
        time.sleep(3)
    
        element = driver_chrome.find_element_by_xpath("/html/body/div[3]/section/aside/nav/ul/li[2]/ul/li[2]/a")
        element.click()
        time.sleep(4)
        























