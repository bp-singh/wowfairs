from selenium.webdriver.common.keys import Keys
from selenium import webdriver
import time
#from selenium.webdriver.support.select import Select 
# import openpyxl module 
import openpyxl 
  
# To open the workbook  
# workbook object is created 
wb_obj = openpyxl.load_workbook("C:\\Users\\BP\\Desktop\\asd.xlsx") 

# Get workbook active sheet object 
# from the active attribute 
sheet_obj = wb_obj.active 

# Cell object is created by using  
# sheet object's cell() method. 

driver_chrome = webdriver.Chrome("C:\chromedriver_win32\chromedriver.exe")
driver_chrome.get("https://wowfairs.com/")
# it will maximize the window size, bydefault the window size is not maximize
driver_chrome.maximize_window()


for i in range(2,6):
    try:
        alpha = sheet_obj.cell(row = i, column = 1)
        a = alpha.value
        b = sheet_obj.cell(row = i, column = 2)
        b = b. value
        c = sheet_obj.cell(row = i, column = 3)
        c = c.value
        d = sheet_obj.cell(row = i, column = 4)
        d = d.value
        e = sheet_obj.cell(row = i, column = 5)
        e = e.value
        f = sheet_obj.cell(row = i, column = 6)
        f = f.value
        g = sheet_obj.cell(row = i, column = 7)
        g = g.value
        h = sheet_obj.cell(row = i, column = 8)
        h = h.value
        
        def click():
            element.click()
            time.sleep(2) 
            
        #search for register butoon    
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/form/div[6]/button[1]")
        click()   # click on register button
        
        # full name
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[1]/input")
        element.send_keys(a)
        time.sleep(2)
        
        # email address
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[2]/input")
        element.send_keys(b)
        
        # password
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[3]/input")
        element.send_keys(c)
        
        # confirm password
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[4]/input")
        element.send_keys(d)
        
        # designation
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[5]/input")
        element.send_keys(e)
        
        # company
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[6]/input")
        element.send_keys(f)
        
        # GST
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[7]/input")
        element.send_keys(g)
        
        #  mobile number
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[8]/ngx-intl-tel-input/div/input")
        element.send_keys(h)
        
        # terms and condition check box
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/div[9]/label/input")
        click()
        '''
        # click on register
        element = driver_chrome.find_element_by_xpath("/html/body/app-root/app-content-layout/div/div/div/app-landing-page/section/div/div[2]/div[2]/app-login/section/div/div/app-register/div/form/button")
        click()
        time.sleep(2)
        '''
        
        
    except:
        driver_chrome.quit()
